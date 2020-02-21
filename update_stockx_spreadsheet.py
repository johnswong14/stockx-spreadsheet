from openpyxl import load_workbook
import requests
import json
import time
import glob, os

# constants
HEADERS = {
        'sec-fetch-mode': 'cors',
        'accept-language': 'en-US,en;q=0.9',
        'authorization': '',
        'x-requested-with': 'XMLHttpRequest',
        'appos': 'web',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36',
        'accept': '*/*',
        'authority': 'stockx.com',
        'sec-fetch-site': 'same-origin',
        'appversion': '0.1',
        "Accept": 'application/json'
        }

# load excel workbook
def load_excel_workbook(filename):
    workbook = load_workbook(filename=filename)
    return workbook
    
# load excel spreadsheet
def load_excel_spreadsheet(workbook):
    workbook.active = 1
    sheet = workbook.active
    return sheet

# retrieve product URL key
def get_URL_key(requested_product):
    try:
        search_URL = f"https://stockx.com/api/browse?&_search={requested_product['title']}"
        source = requests.get(search_URL, headers=HEADERS)
        search_data = source.json()
        
        if not search_data['Products']:
            URL_key = None
            print("No product found")
        else:
            product = search_data['Products'][0]
            URL_key = product['urlKey']
    except:
        URL_key = None
        print("Error has occurred")
        
    return URL_key
    
# retrieve product info (w/ url key)
def get_product_info(product):
    if product['URL_key'] is None:
        print("Missing URL key\n")
    else:
        keyword = product['URL_key'].replace("-", " ")
        URL = f"https://stockx.com/api/browse?&_search={keyword}"
        source = requests.get(URL, headers=HEADERS)
        product_data = source.json()
        
        if not product_data['Products']:
            print("Error fetching product information")
            return None
        else:
            for prod in product_data['Products']:
                if prod['urlKey'] == product['URL_key'] and prod['shoeSize'] == product['size']:
                    title = prod['title'] if prod['title'] else None
                    style_id = prod['styleId'] if prod['styleId'] else None
                    actual_size = prod['shoeSize'] if prod['shoeSize'] else None
                    lowest_ask_price = prod['market']['lowestAsk'] if prod['market']['lowestAsk'] else None
                    highest_bid_price = prod['market']['highestBid'] if prod['market']['highestBid'] else None
                    last_sale_price = prod['market']['lastSale'] if prod['market']['lastSale'] else None

                    product_info = {
                        "title": title, 
                        "style_id": style_id, 
                        "actual_size": actual_size, 
                        "lowest_ask_price": lowest_ask_price, 
                        "highest_bid_price": highest_bid_price, 
                        "last_sale_price": last_sale_price
                    }

                    print(f"Product: {product_info['title']}\nSize: {product_info['actual_size']}\nPrice: {format(product_info['lowest_ask_price'], '.2f')}")
                    return product_info

# update StockX price
def update_price(sheet, start_row, end_row):
    total_items_price_up = 0
    total_items_price_down = 0
    
    for value in sheet.iter_rows(min_row=start_row,
                                 max_row=end_row,
                                 min_col=1,
                                 max_col=12,
                                 values_only=False):
        product = {
            "title": value[0].value,
            "URL_key": value[11].value,
            "size": value[1].value
        }
        
        if (product['URL_key']) is None:
            value[11].value = get_URL_key(product)
            product['URL_key'] = value[11].value
        
        product_info = get_product_info(product)
        
        if product_info is not None:
            previous_price = float(value[5].value)
            
            if product_info['lowest_ask_price']:
                value[5].value = product_info['lowest_ask_price']
                print(f"Lowest Ask Price: {format(value[5].value, '.2f')}")
            elif product_info['last_sale_price']:
                value[5].value = product_info['last_sale_price']
                print(f"Last Sale Price: {format(value[5].value, '.2f')}")
            elif product_info['highest_bid_price']:
                value[5].value = product_info['highest_bid_price']
                print(f"Highest Bid Price: {format(value[5].value, '.2f')}")
                
            print(f"Previous price: {format(previous_price, '.2f')}")    
                
            if value[5].value > previous_price:
                total_items_price_up += 1
                print("PRICE IS BOOMIN'!\n")
            elif value[5].value < previous_price:
                total_items_price_down += 1
                print("Price went down...\n")
            else:
                print("Same old thing.\n")
        
        time.sleep(1)
                
    print(f"Total number of items that went up in price: {total_items_price_up}\nTotal number of items that went down in price: {total_items_price_down}\nTotal number of items that stayed the same in price: {(end_row - start_row + 1) - (total_items_price_down + total_items_price_up)}")

# main program
start_row = int(input("Enter starting row: "))
end_row = int(input("Enter ending row: "))

os.chdir("output")
list_of_files = glob.glob("*.xlsx")
latest_file = max(list_of_files, key=os.path.getctime)
latest_file_version = int(latest_file.split("_")[1].split(".")[0])

print(f"\nLoading the latest file: {latest_file}")
workbook = load_excel_workbook(latest_file)
sheet = load_excel_spreadsheet(workbook)

start = time.time()
print("Updating StockX prices...\n")
update_price(sheet, start_row, end_row)
output_filename = f"balance_{latest_file_version + 1}.xlsx"
workbook.save(filename=output_filename)
end = time.time()
print(f"Finished updating StockX prices in {format(end - start, '.2f')}s")