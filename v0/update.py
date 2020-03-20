import json
import requests
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# constants
FILENAME = "balance_16.xlsx"
HEADERS = {"User-Agent": 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.106 Safari/537.36', "Accept": 'application/json'}

# load excel workbook
def load_excel_workbook(filename):
    workbook = load_workbook(filename=filename)
    return workbook
    
# load excel spreadsheet
def load_excel_spreadsheet(workbook):
    workbook.active = 1
    sheet = workbook.active
    return sheet

# retrieve product URL key
def get_URL_key(requested_product):
    try:
        search_URL = f"https://stockx.com/api/browse?&_search={requested_product['title']}"
        source = requests.get(search_URL, headers=HEADERS)
        search_data = source.json()
        
        if not search_data['Products']:
            URL_key = None
            print("No product found")
        else:
            product = search_data['Products'][0]
            URL_key = product['urlKey']
    except:
        URL_key = None
        print("Error has occurred")
        
    
    return URL_key
    
# convert a string ($1,000) to a float (1000.0)
def convert_to_float(string):
    converted_float = float((string)[1:].replace(",",""))
    return converted_float
    
# retrieve product info (w/ url key)
def get_product_info(product):
    if product['URL_key'] is None:
        print("Missing URL key\n")
    else:
        product_URL = f"https://stockx.com/{product['URL_key']}?size={product['size']}"
        source = requests.get(product_URL, headers=HEADERS)
        product_page = BeautifulSoup(source.content, 'html.parser')
        
        title = product_page.find('h1', {'data-testid': 'product-name'}).text.strip() if product_page.find('h1', {'data-testid': 'product-name'}) else 'error fetching title'
        style_id = product_page.find('span', {'data-testid': 'product-detail-style'}).text.strip() if product_page.find('span', {'data-testid': 'product-detail-style'}) else 'null'
        actual_size = product_page.find('span', {'class': 'variant'}).text.strip() if product_page.find('span', {'class': 'variant'}) else 'OS'
        lowest_ask_price = product_page.findAll('div', {'class': 'stat-value'})[0].text.strip() if product_page.findAll('div', {'class': 'stat-value'}) else None
        highest_bid_price = product_page.findAll('div', {'class': 'stat-value'})[1].text.strip() if product_page.findAll('div', {'class': 'stat-value'}) else None
        last_sale_price = product_page.find('div', {'class': 'sale-value'}).text.strip() if product_page.find('div', {'class': 'sale-value'}) else None
        
        print(lowest_ask_price)
        
        lowest_ask_price = convert_to_float(lowest_ask_price) if lowest_ask_price != "--" or None else lowest_ask_price
        highest_bid_price = convert_to_float(highest_bid_price) if highest_bid_price != "--" or None else highest_bid_price
        last_sale_price = convert_to_float(last_sale_price) if last_sale_price != "--" or None else last_sale_price

        """ data = f"Product Name: {product_title}\nStyle ID: {product_style_id}\nSize: {actual_size}\nLowest Ask Price: {lowest_ask_price}\nHighest Bid Price: {highest_bid_price}\nLast Sale Price: {last_sale_price}\n"

        print(data) """

        product_info = {
            "title": title, 
            "style_id": style_id, 
            "actual_size": actual_size, 
            "lowest_ask_price": lowest_ask_price, 
            "highest_bid_price": highest_bid_price, 
            "last_sale_price": last_sale_price
        }

        print(f"Product: {product_info['title']}\nSize: {product_info['actual_size']}\nPrice: {product_info['lowest_ask_price']}")
        return product_info

# update StockX price
def update_price(sheet, start_row, end_row):
    total_items_price_up = 0
    total_items_price_down = 0
    
    for value in sheet.iter_rows(min_row=start_row,
                                 max_row=end_row,
                                 min_col=1,
                                 max_col=12,
                                 values_only=False):
        product = {
            "title": value[0].value,
            "URL_key": value[11].value,
            "size": value[1].value
        }
        
        if (product['URL_key']) is None:
            value[11].value = get_URL_key(product)
            product['URL_key'] = get_URL_key(product)
        
        product_info = get_product_info(product)
        
        if product_info is not None:
            previous_price = float(value[5].value)
            
            if product_info['lowest_ask_price'] != "--":
                value[5].value = product_info['lowest_ask_price']
                print(f"Lowest Ask Price: {format(value[5].value, '.2f')}")
            elif product_info['last_sale_price'] != "--":
                value[5].value = product_info['last_sale_price']
                print(f"Last Sale Price: {format(value[5].value, '.2f')}")
            elif product_info['highest_bid_price'] != "--":
                value[5].value = product_info['highest_bid_price']
                print(f"Highest Bid Price: {format(value[5].value, '.2f')}")
                
            print(f"Previous price: {format(previous_price, '.2f')}")    
                
            if value[5].value > previous_price:
                total_items_price_up += 1
                print("PRICE IS BOOMIN'!\n")
            elif value[5].value < previous_price:
                total_items_price_down += 1
                print("Price went down...\n")
            else:
                print("Same old thing.\n")
        
        time.sleep(1)
                
    print(f"Total number of items that went up in price: {total_items_price_up}\nTotal number of items that went down in price: {total_items_price_down}\nTotal number of items that stayed the same in price: {(end_row - start_row + 1) - (total_items_price_down + total_items_price_up)}")

# main program
start_row = 123
end_row = 124
start = time.time()
print("Updating StockX prices...\n")
workbook = load_excel_workbook(FILENAME)
sheet = load_excel_spreadsheet(workbook)
#get_url_key(sheet, start_row, end_row)
update_price(sheet, start_row, end_row)
workbook.save(filename="balance_17.xlsx")
end = time.time()
print(f"Finished updating StockX prices in {format(end - start, '.2f')}s")