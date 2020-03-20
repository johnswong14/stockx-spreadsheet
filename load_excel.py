import config
from openpyxl import load_workbook

def load_excel_workbook(filename):
    try:
        workbook = load_workbook(filename=filename)
        return workbook
    except:
        print("Failed loading Excel workbook")

def load_excel_spreadsheet(workbook):
    try:
        workbook.active = config.ACTIVE_SPREADSHEET
        sheet = workbook.active
        return sheet
    except:
        print("Failed loading Excel spreadsheet")

def is_format_valid(sheet):
    if (sheet['A1'].value == "Item" 
            and sheet['B1'].value == "Size" 
            and sheet['C1'].value == "StockX Ticker Symbol" 
            and sheet['D1'].value == "Price Paid" 
            and sheet['G1'].value == "StockX Price"):
        return True
    return False

def find_end_row(sheet):
    try:
        end_row = sheet.max_row
    
        while sheet.cell(column=1, row=end_row).value is None and end_row > 0:
            end_row -= 1
        end_row -=1
        
        return end_row
    except:
        print("Failed finding end row in spreadsheet")